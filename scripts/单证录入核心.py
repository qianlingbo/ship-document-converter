#!/usr/bin/env python3
"""
单证录入核心脚本
读取原始 crew list + port of call 文件，按规则生成标准格式录入文件
"""

import os
import re
import sys
import json
import random
import warnings
from pathlib import Path
from datetime import datetime, timedelta

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

warnings.filterwarnings('ignore')

# ── 路径配置 ──────────────────────────────────────────────────────────────
WORKSPACE = Path(__file__).parent.parent.resolve()
REF_DIR   = WORKSPACE / "references"
INPUT_DIR = WORKSPACE / "input"
OUTPUT_DIR = WORKSPACE / "output"
TEMPLATE_PATH = WORKSPACE / "templates" / "单证录入标准格式.xlsx"

# ── 加载参数映射 ────────────────────────────────────────────────────────────
def load_refs():
    with open(REF_DIR / "nationality_map.json", encoding="utf-8") as f:
        nat_map = json.load(f)
    with open(REF_DIR / "duty_map.json", encoding="utf-8") as f:
        duty_list = list(json.load(f).keys())
    with open(REF_DIR / "port_map.json", encoding="utf-8") as f:
        port_map = json.load(f)
    return nat_map, duty_list, port_map

NATIONALITY_MAP, DUTY_LIST, PORT_MAP = load_refs()

# ── 工具函数 ────────────────────────────────────────────────────────────────

def normalize_code(val, mapping):
    """模糊匹配：输入值 → 参数代码（如 GR→GR-希腊，VIETNAM→VN-越南）"""
    if not val:
        return None
    v = str(val).strip().upper()

    # 1. 精确匹配：code / full / full-split-0
    for code, full in mapping.items():
        if v == code.upper() or v == full.upper() or v == full.split("-")[0].upper():
            return full

    # 2. 部分匹配：v in full or full in v
    for code, full in mapping.items():
        if v in full.upper() or full.upper() in v:
            return full

    # 3. 国籍全称兜底（raw数据可能是全称如VIETNAM/MYANMAR）
    # 从full值中提取国家名（取"-"后面的中文/英文部分）
    COUNTRY_NAME_TO_CODE = {
        "VIETNAM": "VN", "VIET NAM": "VN", "VIETNAMESE": "VN",
        "MYANMAR": "MM", "BURMA": "MM",
        "INDONESIA": "ID", "INDONESIAN": "ID",
        "CHINA": "CN", "CHINESE": "CN",
    }
    # 提取映射中的英文国家名（"-英文名" 部分）
    for code, full in mapping.items():
        parts = full.split("-", 1)
        if len(parts) == 2:
            en_name = parts[1].strip()
            # 去掉括号中的内容
            if "(" in en_name:
                en_name = en_name.split("(")[0].strip()
            if en_name and en_name.isalpha():
                COUNTRY_NAME_TO_CODE[en_name.upper()] = code

    raw_name = v.replace(",", " ").replace(".", " ").strip()
    for cname, ccode in COUNTRY_NAME_TO_CODE.items():
        if raw_name == cname or raw_name.startswith(cname) or cname.startswith(raw_name):
            if ccode in mapping:
                return mapping[ccode]

    # 4. 英文国家名直译兜底
    ENGLISH_COUNTRY_NAMES = {
        "CHINA": "CN", "CHINESE": "CN",
        "VIETNAM": "VN", "VIETNAMESE": "VN", "VIET NAM": "VN",
        "MYANMAR": "MM", "BURMESE": "MM",
        "INDONESIA": "ID", "INDONESIAN": "ID",
        "PANAMA": "PA",
        "KOREA": "KR", "SOUTH KOREA": "KR", "REPUBLIC OF KOREA": "KR",
        "JAPAN": "JP", "JAPANESE": "JP",
        "HONGKONG": "HK", "HONG KONG": "HK",
        "MACAO": "MO", "MACAU": "MO",
        "RUSSIA": "RU", "RUSSIAN": "RU",
        "UNITED STATES": "US", "USA": "US", "AMERICA": "US",
        "UNITED KINGDOM": "GB", "UK": "GB", "BRITAIN": "GB",
        "GERMANY": "DE", "GERMAN": "DE",
        "PHILIPPINES": "PH",
        "THAILAND": "TH",
        "MALAYSIA": "MY",
        "SINGAPORE": "SG",
        "NORWAY": "NO",
        "GREECE": "GR",
        "MARSHALL ISLANDS": "MH",
    }
    for en_name, ccode in ENGLISH_COUNTRY_NAMES.items():
        if raw_name == en_name or raw_name.startswith(en_name):
            if ccode in mapping:
                return mapping[ccode]

    return None

def match_port(val):
    """匹配港口：输入值 → 参数E列完整字符串"""
    if not val:
        return None
    v = str(val).strip().upper()

    # 1. 标准匹配
    for code, full in PORT_MAP.items():
        if v == code.upper() or v == full.upper() or v == full.split("-")[0].upper():
            return full
    for code, full in PORT_MAP.items():
        if v in full.upper() or full.upper().replace(" ", "") in v.replace(" ", ""):
            return full

    # 2. 拆分组合字符串，如 "LIANYUNGANG, CHINA" → 取第一段匹配
    if "," in v or "-" in v:
        parts = [p.strip() for p in v.replace("-", ",").split(",")]
        for part in parts:
            if len(part) > 3:  # 排除 CHINA 等太短的国家名
                for code, full in PORT_MAP.items():
                    if part == code.upper() or part in full.upper() or full.upper() in part:
                        return full
                    # 部分匹配：LIANYUNGANG → 连云港（要求 ≥6 字符防误匹配）
                    if (len(part) >= 6 and part[:6] in full.upper()) or \
                       (len(full.split("-")[0]) >= 6 and full.split("-")[0][:6] in part.upper()):
                        return full

    return None

def random_port_same_country(port_raw):
    """港口匹配不到时，根据港口代码前两位（国家代码）随机选择同国家的其他港口"""
    if not port_raw:
        return None
    v = str(port_raw).strip().upper()
    # 尝试提取国家代码：港口代码前两位，或逗号后面的国家名
    country_prefix = None
    if len(v) >= 2 and v[:2].isalpha():
        country_prefix = v[:2]
    elif "," in v:
        parts = [p.strip() for p in v.split(",")]
        # 最后一段可能是国家名，从 NATIONALITY_MAP 里找代码
        for part in reversed(parts):
            code = normalize_code(part, NATIONALITY_MAP)
            if code and len(code) >= 2:
                # 取代码前两位
                country_prefix = code[:2]
                break
    
    if not country_prefix:
        return None
    
    # 找所有同国家前缀的港口
    candidates = [full for code, full in PORT_MAP.items() if code[:2].upper() == country_prefix]
    if candidates:
        return random.choice(candidates)
    return None

def match_duty(val):
    """匹配职务：输入值 → 参数B列完整字符串，找不到返回默认值规则"""
    if not val:
        return None
    v = str(val).strip().upper()

    # 英语缩写 → 标准职务
    ENGLISH_RANK_MAP = {
        "MASTER": "51-船长", "CAPT": "51-船长", "CAPTAIN": "51-船长",
        "C/O": "52-大副", "CHIEF OFFICER": "52-大副", "FIRST OFFICER": "52-大副",
        "2/O": "53-二副", "SECOND OFFICER": "53-二副",
        "3/O": "54-三副", "THIRD OFFICER": "54-三副",
        "C/E": "61-轮机长", "CHIEF ENGINEER": "61-轮机长",
        "2/E": "62-大管轮", "SECOND ENGINEER": "62-大管轮",
        "3/E": "63-二管轮", "THIRD ENGINEER": "63-二管轮",
        "4/E": "64-三管轮", "FOURTH ENGINEER": "64-三管轮",
        "BSN": "55-值班水手", "BOSUN": "55-值班水手",
        "AB1": "56-高级值班水手", "ABLE SEAMAN": "56-高级值班水手",
        "AB2": "55-值班水手",
        "AB3": "55-值班水手",
        "D/C": "65-值班机工", "DECK CADET": "65-值班机工",
        "OIL1": "66-高级值班机工", "OIL": "66-高级值班机工",
        "OIL2": "66-高级值班机工",
        "OIL3": "66-高级值班机工",
        "FITTER": "65-值班机工", "FTR": "65-值班机工",
        "E/E": "66-高级值班机工", "ELECTRICIAN": "66-高级值班机工",
        "ETO": "66-高级值班机工", "ELECTRO-TECHNICAL OFFICER": "66-高级值班机工",
        "OLR": "66-高级值班机工", "OILER": "66-高级值班机工",
        "AB4": "55-值班水手",
        "COOK": "65-值班机工", "CHEF": "65-值班机工", "GALLEY": "65-值班机工",
        "C/CK": "65-值班机工", "CHIEF COOK": "65-值班机工",
        "D/T": "55-值班水手", "DECK TRAINEE": "55-值班水手",
        "E/T": "65-值班机工", "ENGINE TRAINEE": "65-值班机工",
        "PUMPMAN": "65-值班机工", "P/MAN": "65-值班机工",
        "STEWARD": "65-值班机工", "STD": "65-值班机工",
        "MESSMAN": "65-值班机工", "M/M": "65-值班机工",
        "TALLY CLERK": "65-值班机工",
        "AB": "56-高级值班水手", "A/B": "56-高级值班水手",
        "OS": "55-值班水手", "ORDINARY SEAMAN": "55-值班水手",
        "WPR": "65-值班机工", "WIPER": "65-值班机工",
        "1/E": "62-大管轮", "FIRST ENGINEER": "62-大管轮",
        "1/O": "52-大副",
        "RADIO OFFICER": "66-高级值班机工", "R/O": "66-高级值班机工",
        "CADET": "55-值班水手",
        "TRAINEE": "55-值班水手",
        "S/M": "65-值班机工",
    }
    if v in ENGLISH_RANK_MAP:
        return ENGLISH_RANK_MAP[v]

    for duty in DUTY_LIST:
        if v == duty.upper() or v == duty.split("-")[1].upper().strip():
            return duty
    for duty in DUTY_LIST:
        if v in duty.upper() or duty.upper() in v:
            return duty
    return None  # 由调用方按规则分配

def normalize_date(val):
    """各种日期格式 → YYYYMMDD"""
    if not val:
        return None
    # 支持 datetime 对象
    if hasattr(val, 'strftime'):
        return val.strftime('%Y%m%d')
    s = str(val).strip().replace(" ", "")
    for fmt in ["%Y-%m-%d", "%Y/%m/%d", "%Y%m%d", "%d/%m/%Y", "%d-%m-%Y",
                "%d/%b/%Y", "%d-%b-%Y", "%d/%B/%Y", "%d-%B-%Y"]:
        try:
            return datetime.strptime(s, fmt).strftime("%Y%m%d")
        except:
            pass
    # 去掉多余字符再试
    s2 = s.replace(" ", "").replace(".", "")
    for fmt in ["%Y%m%d"]:
        try:
            return datetime.strptime(s2[:8], fmt).strftime("%Y%m%d")
        except:
            pass
    return None

def random_time_0000_1200():
    h = random.randint(0, 11)
    m = random.randint(0, 59)
    s = random.randint(0, 59)
    return f"{h:02d}:{m:02d}:{s:02d}"

def random_time_1200_2400():
    h = random.randint(12, 23)
    m = random.randint(0, 59)
    s = random.randint(0, 59)
    return f"{h:02d}:{m:02d}:{s:02d}"

def get_nationality_chinese(code):
    """国籍代码 → 中文名（用于出生地点）"""
    mapped = normalize_code(code, NATIONALITY_MAP)
    if mapped and "-" in mapped:
        return mapped.split("-", 1)[1]
    return mapped or code or ""

def get_country_name_for_port(port_val):
    """港口 → 国家/地区名称（参数A格式）"""
    if not port_val:
        return None
    v = str(port_val).upper()

    # 1. 从原始值提取国家部分（如 "HITACHINAKA, JAPAN" → "JAPAN"）
    #    规则：取逗号分隔后的最后一段（≥3字符）直接 normalize
    if "," in v or "-" in v:
        parts = [p.strip() for p in v.replace("-", ",").split(",")]
        for part in reversed(parts):
            if len(part) >= 3:
                # 跳过太长（不是国家名）的部分
                if len(part) > 20:
                    continue
                result = _normalize_country_name(part, NATIONALITY_MAP)
                if result:
                    return result

    # 2. 从港口映射码提取国家码（如 JPHIC → JP → 日本）
    matched = match_port(port_val)
    if matched:
        code = matched.split("-")[0]
        if len(code) <= 4:  # 保护：港口码长度有限
            result = normalize_code(code, NATIONALITY_MAP)
            if result:
                return result

    return None


def _normalize_country_name(raw_name, mapping):
    """国家英文名 → 参数A格式，跳过逗号分隔的国家名"""
    v = str(raw_name).strip().upper()
    # 跳过组合字符串（包含逗号说明是完整地址）
    if "," in v:
        return None
    # 跳过太长的值（正常国家名不会超过20字符）
    if len(v) > 20:
        return None
    return normalize_code(v, mapping)

def _extract_chinese(text):
    """从混合文本中提取中文字符部分（保留中文及中间空格）"""
    chars = re.findall(r'[\u4e00-\u9fff\u3400-\u4dbf]', text)
    return ''.join(chars) if chars else None

def _has_chinese(text):
    """检测文本中是否包含中文字符"""
    return bool(re.search(r'[\u4e00-\u9fff\u3400-\u4dbf]', text))

# ── 职务分配规则 ──────────────────────────────────────────────────────────
# 找不到B列时：保证3个高级值班水手(56) + 3个高级值班机工(66)
# 其余分配55-值班水手和65-值班机工

def assign_duty_fallback(crew_list):
    """对未匹配职务的船员，按规则分配值班水手/机工"""
    unmatched = [c for c in crew_list if c.get("_duty_fallback")]
    # 先填56-高级值班水手（最多3个）
    sailors = [c for c in unmatched if c.get("_role_type") == "sailor"]
    engineers = [c for c in unmatched if c.get("_role_type") == "engineer"]
    
    # 分配高级值班水手（56），最多3个
    for c in sailors[:3]:
        c["船员职务"] = "56-高级值班水手"
    # 分配高级值班机工（66），最多3个
    for c in engineers[:3]:
        c["船员职务"] = "66-高级值班机工"
    # 剩余的分配55和65
    for c in sailors[3:]:
        c["船员职务"] = "55-值班水手"
    for c in engineers[3:]:
        c["船员职务"] = "65-值班机工"

# ── 从Excel读取crew list ──────────────────────────────────────────────────
def _find_header_row(ws, keywords, require_all=True):
    """
    扫描所有行，找到包含指定关键词的行作为表头。
    require_all=True:  必须同时包含所有关键词（适用于"Voy."+ "Port"组合）
    require_all=False:  包含任意关键词即可
    """
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        row_text = " ".join(str(v).upper() for v in row if v is not None)
        if require_all:
            if all(k.upper() in row_text for k in keywords):
                return i, [str(h).strip() if h else "" for h in row]
        else:
            if any(k.upper() in row_text for k in keywords):
                return i, [str(h).strip() if h else "" for h in row]
    return None, None

def _parse_combined_field(val):
    """解析IMO Crew List中日期+文本的组合字段，如 '28/Nov/1986  SHANDONG' → (date_str, text_str)"""
    if not val:
        return None, None
    s = str(val).strip()
    # 尝试用正则拆分：日期部分 (dd/Mon/yyyy 或 dd/mm/yyyy) + 剩余文本
    # 匹配 dd/Mon/yyyy 或 dd/MM/yyyy 格式的日期
    m = re.match(r'(\d{1,2}[/-]\s*\w{2,9}[/-]\s*\d{4})\s*(.*)', s)
    if m:
        return m.group(1).replace(' ', ''), m.group(2).strip()
    # 匹配 yyyy/mm/dd 格式
    m = re.match(r'(\d{4}[/-]\d{1,2}[/-]\d{1,2})\s*(.*)', s)
    if m:
        return m.group(1).strip(), m.group(2).strip()
    return s, None

def read_crew_excel(path):
    """读取任意格式的crew list Excel，返回标准化数据列表"""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    # 扫描定位表头行（关键词：No. + Family name + Rank）
    header_idx, headers = _find_header_row(ws, ["No.", "Family name", "Rank"])
    if header_idx is None:
        # 回退：尝试只匹配 No. + Rank
        header_idx, headers = _find_header_row(ws, ["No.", "Rank"])
    if header_idx is None:
        # 回退：用第1行
        header_idx = 0
        all_rows = list(ws.iter_rows(values_only=True))
        headers = [str(h).strip() if h else "" for h in all_rows[0]]
    print(f"  表头行={header_idx+1}, 列数={len(headers)}: {[h for h in headers if h][:12]}")

    # 构建列名→索引的映射（模糊匹配）
    col_map = {}
    for idx, h in enumerate(headers):
        hu = h.upper()
        if "NO" in hu and ("NO" == hu.strip().rstrip(".") or "NO." in hu):
            col_map.setdefault("no", idx)
        if "NAME" in hu or "FAMILY" in hu:
            col_map.setdefault("name", idx)
        if "RANK" in hu or "RATING" in hu:
            col_map.setdefault("rank", idx)
        if "SEX" in hu:
            col_map.setdefault("sex", idx)
        if "NATIONAL" in hu:
            col_map.setdefault("nation", idx)
        if "BIRTH" in hu:
            col_map.setdefault("birth", idx)
        if "SEAMAN" in hu or "BOOK" in hu:
            col_map.setdefault("seaman_book", idx)
        if "PASSPORT" in hu:
            col_map.setdefault("passport", idx)
        if "JOIN" in hu or ("DATE" in hu and "PLACE" in hu and "JOIN" in hu):
            col_map.setdefault("join", idx)
    print(f"  列映射: {col_map}")

    crew_data = []

    # 扫描所有行，找到数据区（序号列含整数的行）
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i <= header_idx:
            continue
        if not any(v is not None for v in row):
            continue

        def g(j): return row[j] if j is not None and j < len(row) else None

        no_val = g(col_map.get("no", 0))
        name = g(col_map.get("name", 1))
        rank = g(col_map.get("rank", 2))
        sex = g(col_map.get("sex", 3))
        nation = g(col_map.get("nation", 4))
        birth_combined = g(col_map.get("birth", 5))
        seaman_combined = g(col_map.get("seaman_book", 6))
        passport_combined = g(col_map.get("passport", 7))
        join_combined = g(col_map.get("join", 8))

        # 只保留有姓名的行
        if not name:
            continue

        # 序号必须是整数（支持int、float、纯数字字符串）
        if isinstance(no_val, (int, float)) or (isinstance(no_val, str) and no_val.strip().isdigit()):
            # 解析组合字段：日期 + 地点/编号
            birth_date, birth_place = _parse_combined_field(birth_combined)
            seaman_exp, seaman_no = _parse_combined_field(seaman_combined)
            passport_exp, passport_no = _parse_combined_field(passport_combined)
            join_date, join_place = _parse_combined_field(join_combined)

            c = {
                "_raw_name": name,
                "_raw_sex": sex,
                "_raw_duty": rank,
                "_raw_nation": nation,
                "_raw_birth": birth_date,
                "_raw_birth_place": birth_place,
                "_raw_passport": passport_no,
                "_raw_passport_exp": passport_exp,
                "_raw_seaman_no": seaman_no,
                "_raw_seaman_exp": seaman_exp,
                "_raw_port": join_place,
                "_raw_joindate": join_date,
            }
            crew_data.append(c)

    return crew_data

# ── 从PDF读取crew list ─────────────────────────────────────────────────────
def read_crew_pdf(path):
    """用 pdfplumber 读取 PDF，返回标准化数据列表"""
    try:
        import pdfplumber
    except ImportError:
        print("请安装 pdfplumber: pip install pdfplumber")
        return []
    
    crew_data = []
    with pdfplumber.open(path) as pdf:
        for page in pdf.pages:
            tables = page.extract_tables()
            for table in tables:
                for row in table:
                    if not row or not any(row):
                        continue
                    c = {}
                    for j, val in enumerate(row):
                        if not val:
                            continue
                        vs = str(val).strip()
                        # 尝试各字段
                        # 这里需要根据实际PDF布局调整
                        if len(vs) > 2 and vs.replace(" ", "").isalpha():
                            c["_raw_name"] = val
                        elif "/" in vs or "-" in vs:
                            if len(vs) == 10 or len(vs) == 8:
                                c["_raw_birth"] = val
                    if c.get("_raw_name"):
                        crew_data.append(c)
    
    return crew_data

# ── 从Excel读取port of call ───────────────────────────────────────────────
def read_port_excel(path):
    """读取 port of call Excel"""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    # 扫描定位表头行（必须同时含"Voy."和"Port"）
    header_idx, headers = _find_header_row(ws, ["Voy.", "Port"])
    if header_idx is None:
        header_idx = 0
        all_rows = list(ws.iter_rows(values_only=True))
        headers = [str(h).strip() if h else "" for h in all_rows[0]]
    print(f"  表头行={header_idx+1}, 列数={len(headers)}: {[h for h in headers if h][:10]}")

    ports_data = []

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i <= header_idx:
            continue
        if not any(v is not None for v in row):
            continue

        def g(j): return row[j] if j < len(row) else None

        # col B(1)=Voy.No, C(2)=Port, D(3)=Arrival, E(4)=Departure,
        # F(5)=Cargo, G(6)=SecurityLevel(Port), H(7)=SecurityLevel(Ship),
        # I(8)=Additional Security
        port = g(2)
        arrival = g(3)
        departure = g(4)

        # 只保留有港口名的行
        if not port:
            continue

        p = {
            "_raw_port": port,
            "_raw_arrival": arrival,
            "_raw_departure": departure,
        }
        ports_data.append(p)

    return ports_data

# ── 从PDF读取port of call ──────────────────────────────────────────────────
def read_port_pdf(path):
    """用 pdfplumber 读取 port of call PDF"""
    try:
        import pdfplumber
    except ImportError:
        print("请安装 pdfplumber: pip install pdfplumber")
        return []
    
    ports_data = []
    with pdfplumber.open(path) as pdf:
        for page in pdf.pages:
            tables = page.extract_tables()
            for table in tables:
                for row in table:
                    if not row or not any(row):
                        continue
                    p = {}
                    for j, val in enumerate(row):
                        if not val:
                            continue
                        vs = str(val).strip()
                        if "PORT" in vs.upper() or any(c.isalpha() for c in vs):
                            if len(vs) > 2:
                                p["_raw_port"] = val
                    if p:
                        ports_data.append(p)
    
    return ports_data

# ── 船员数据标准化 ─────────────────────────────────────────────────────────
def normalize_crew(raw_list, default_port=None, default_joindate=None):
    """
    将原始船员数据按规则标准化
    default_port: 默认登船口岸（从port of call第一个港口）
    default_joindate: 默认登船日期
    """
    result = []
    
    for idx, c in enumerate(raw_list, 1):
        # 国籍
        nation_code = c.get("_raw_nation", "")
        nation_mapped = normalize_code(nation_code, NATIONALITY_MAP)
        _nation_fallback = False
        if not nation_mapped:
            nation_mapped = ""  # 匹配不到留空
            _nation_fallback = True
        
        nation_code2 = nation_mapped.split("-")[0] if nation_mapped else "CN"
        is_chinese = nation_code2 == "CN"
        
        # 姓名
        name = str(c.get("_raw_name", "")).strip()
        if not name:
            continue
        if is_chinese:
            # 中国船员：只保留中文部分（如果原始数据同时有英文和中文）
            if _has_chinese(name):
                chinese_part = _extract_chinese(name)
                if chinese_part:
                    name = chinese_part
            # 如果只有英文（没有中文字符），保留原样
        else:
            # 外国船员：全部转大写英文
            name = name.upper()
        
        # 性别
        sex_raw = c.get("_raw_sex", "")
        sex_map = {"M": "1-男", "F": "2-女", "男": "1-男", "女": "2-女",
                   "1": "1-男", "2": "2-女", "male": "1-男", "female": "2-女"}
        sex = sex_map.get(str(sex_raw).strip().upper(), "1-男")
        
        # 出生日期
        birth = normalize_date(c.get("_raw_birth", ""))
        
        # 出生地点
        # 中国船员：无论源数据写了什么（省市、拼音、英文），一律输出"中国"
        # 外国船员：优先保留源数据原文，无源数据时用国籍中文名兜底
        if is_chinese:
            birth_place = "中国"
        else:
            raw_bp = str(c.get("_raw_birth_place", "") or "").strip()
            birth_place = raw_bp if raw_bp else get_nationality_chinese(nation_code2)
        
        # 职务
        duty_raw = c.get("_raw_duty", "")
        duty_mapped = match_duty(duty_raw)
        duty_fallback = False
        role_type = None
        if not duty_mapped:
            duty_fallback = True
            # 尝试识别是水手还是机工
            duty_upper = duty_raw.upper()
            if any(w in duty_upper for w in ["ENGINE", "MECH", "机", "轮", "电机", "管"]):
                role_type = "engineer"
            else:
                role_type = "sailor"
            duty_mapped = ""  # 待分配
        
        # 证件类型 & 证件号码
        if is_chinese:
            cert_type = "17-海员证"
            # 中国船员优先使用海员证号码
            seaman_no = str(c.get("_raw_seaman_no", "")).strip()
            passport = seaman_no if seaman_no else str(c.get("_raw_passport", "")).strip()
        else:
            cert_type = "14-普通护照"
            passport = str(c.get("_raw_passport", "")).strip()
        
        # 登船口岸
        port_raw = c.get("_raw_port", default_port)
        port_mapped = match_port(port_raw) if port_raw else None
        if not port_mapped and default_port:
            port_mapped = match_port(default_port)
        
        # 登船日期
        joindate_raw = c.get("_raw_joindate", default_joindate)
        joindate = normalize_date(joindate_raw)
        if not joindate and default_joindate:
            joindate = normalize_date(default_joindate)
        
        entry = {
            "序号": str(idx),
            "姓名": name,
            "性别": sex,
            "船员职务": duty_mapped or "",
            "_duty_fallback": duty_fallback,
            "_role_type": role_type,
            "船员国籍 ": nation_mapped or "",
            "出生日期": birth or "",
            "出生地点": birth_place,
            "证件类型": cert_type,
            "证件号码": passport,
            "是否申请登陆 ": "",      # 空白
            "适任证书编号 ": "",      # 空白
            "适任证书有效期至": "",    # 空白
            "证件检查地点": "",        # 空白
            "登船日期": joindate or "",
            "登船口岸": port_mapped or "",
            "备注": "",
            "_nation_fallback": _nation_fallback
        }
        result.append(entry)
    
    # 处理职务未匹配的船员
    assign_duty_fallback(result)
    
    return result

# ── 港口数据标准化 ─────────────────────────────────────────────────────────
def normalize_ports(raw_list):
    """将原始港口数据按规则标准化"""
    result = []
    
    for idx, p in enumerate(raw_list, 1):
        port_raw = p.get("_raw_port", "")
        country_raw = p.get("_raw_country", "")
        arrival_raw = p.get("_raw_arrival", "")
        departure_raw = p.get("_raw_departure", "")
        
        # 港口
        port_mapped = match_port(port_raw)
        _port_fallback = False
        if not port_mapped:
            # 匹配不到：随机选同国家的其他港口，标红
            port_mapped = random_port_same_country(port_raw)
            if port_mapped:
                _port_fallback = True
        
        # 国家/地区
        if country_raw:
            country_mapped = normalize_code(country_raw, NATIONALITY_MAP)
        else:
            country_mapped = get_country_name_for_port(port_raw)
        
        # 进港时间：日期 + 0000-1200随机时间
        arrival_date = normalize_date(arrival_raw)
        if arrival_date:
            arrival_date_fmt = f"{arrival_date[:4]}/{arrival_date[4:6]}/{arrival_date[6:8]}"
            arrival_time = random_time_0000_1200()
            arrival_full = f"{arrival_date_fmt} {arrival_time}"
        else:
            arrival_full = ""
        
        # 离港时间：日期 + 1200-2400随机时间
        departure_date = normalize_date(departure_raw or arrival_raw)
        if departure_date:
            departure_date_fmt = f"{departure_date[:4]}/{departure_date[4:6]}/{departure_date[6:8]}"
            departure_time = random_time_1200_2400()
            departure_full = f"{departure_date_fmt} {departure_time}"
        else:
            departure_full = ""
        
        # 保安等级
        security_level = "1-1级"
        port_security = "1-1级"
        
        entry = {
            "序号": idx,
            "进港时间": arrival_full,
            "离港时间": departure_full,
            "国家/地区名称": country_mapped or "",
            "船舶保安等级": security_level,
            "特别或附加的保安设施": "",  # 空白
            "停靠港口": port_mapped or "",
            "港口保安等级": port_security,
            "_port_fallback": _port_fallback
        }
        result.append(entry)
    
    return result

# ── 写入Excel ──────────────────────────────────────────────────────────────
RED_FILL = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")

def _mark_row_red(ws, row, max_col):
    """将整行标红（浅红背景）"""
    for col in range(1, max_col + 1):
        ws.cell(row=row, column=col).fill = RED_FILL

def fill_crew_sheet(template_wb, crew_list):
    """填写船员名单sheet"""
    ws = template_wb["船上非旅客人员清单"]
    # 找到数据起始行（第3行）
    start_row = 3
    # 先清空旧数据（保留前两行格式行）
    for row in range(start_row, ws.max_row + 50):
        for col in range(1, 17):
            ws.cell(row=row, column=col).value = None
    
    for i, c in enumerate(crew_list):
        row = start_row + i
        ws.cell(row=row, column=1).value = c["序号"]
        ws.cell(row=row, column=2).value = c["姓名"]
        ws.cell(row=row, column=3).value = c["性别"]
        ws.cell(row=row, column=4).value = c["船员职务"]
        ws.cell(row=row, column=5).value = c["船员国籍 "]
        ws.cell(row=row, column=6).value = c["出生日期"]
        ws.cell(row=row, column=7).value = c["出生地点"]
        ws.cell(row=row, column=8).value = c["证件类型"]
        ws.cell(row=row, column=9).value = c["证件号码"]
        ws.cell(row=row, column=10).value = c["是否申请登陆 "]
        ws.cell(row=row, column=11).value = c["适任证书编号 "]
        ws.cell(row=row, column=12).value = c["适任证书有效期至"]
        ws.cell(row=row, column=13).value = c["证件检查地点"]
        ws.cell(row=row, column=14).value = c["登船日期"]
        ws.cell(row=row, column=15).value = c["登船口岸"]
        
        # 国籍匹配不到 → 整行标红
        if c.get("_nation_fallback"):
            _mark_row_red(ws, row, 16)
        ws.cell(row=row, column=16).value = c["备注"]

def fill_goods_sheet(template_wb, crew_list):
    """填写物品清单sheet：证件类型/号码参照船员名单，其余固定"""
    ws = template_wb["船上非旅客人员物品清单"]
    start_row = 3
    
    # 清空旧数据
    for row in range(start_row, ws.max_row + 50):
        for col in range(1, 11):
            ws.cell(row=row, column=col).value = None
    
    for i, c in enumerate(crew_list):
        row = start_row + i
        ws.cell(row=row, column=1).value = str(i + 1)         # 序号
        ws.cell(row=row, column=2).value = c["证件类型"]        # 证件类型
        ws.cell(row=row, column=3).value = c["证件号码"]        # 证件号码
        ws.cell(row=row, column=4).value = "0100"              # 物品类型：计算机
        ws.cell(row=row, column=5).value = "计算机"             # 物品名称
        ws.cell(row=row, column=6).value = 1                   # 物品数量
        ws.cell(row=row, column=7).value = "001"              # 数量单位

def fill_port_sheet(template_wb, ports_list):
    """填写海事船岸活动信息sheet"""
    ws = template_wb["海事船岸活动信息"]
    start_row = 3
    
    # 清空旧数据
    for row in range(start_row, ws.max_row + 50):
        for col in range(1, 9):
            ws.cell(row=row, column=col).value = None
    
    for i, p in enumerate(ports_list):
        row = start_row + i
        ws.cell(row=row, column=1).value = p["序号"]
        ws.cell(row=row, column=2).value = p["进港时间"]
        ws.cell(row=row, column=3).value = p["离港时间"]
        ws.cell(row=row, column=4).value = p["国家/地区名称"]
        ws.cell(row=row, column=5).value = p["船舶保安等级"]
        ws.cell(row=row, column=6).value = p["特别或附加的保安设施"]
        ws.cell(row=row, column=7).value = p["停靠港口"]
        ws.cell(row=row, column=8).value = p["港口保安等级"]
        
        # 停靠港口匹配不到、用了同国家随机港口 → 整行标红
        if p.get("_port_fallback"):
            _mark_row_red(ws, row, 8)

# ── 主入口 ─────────────────────────────────────────────────────────────────
def process(crew_path, port_path=None, output_name=None):
    """
    主处理函数
    crew_path: crew list 文件路径 (Excel 或 PDF)
    port_path: port of call 文件路径 (Excel 或 PDF)，可选
    output_name: 输出文件名（不含扩展名）
    """
    print(f"\n{'='*60}")
    print(f"  单证录入处理")
    print(f"{'='*60}")
    
    crew_path = Path(crew_path)
    if not crew_path.exists():
        print(f"❌ 文件不存在: {crew_path}")
        return None
    
    # 1. 读取船员数据
    print(f"\n[1/4] 读取船员数据: {crew_path.name}")
    if crew_path.suffix.lower() in [".xlsx", ".xls"]:
        raw_crew = read_crew_excel(crew_path)
    elif crew_path.suffix.lower() == ".pdf":
        raw_crew = read_crew_pdf(crew_path)
    else:
        print(f"❌ 不支持的船员文件格式: {crew_path.suffix}")
        return None
    print(f"  → 读取到 {len(raw_crew)} 条原始记录")
    
    # 2. 读取港口数据
    default_port = None
    default_joindate = None
    raw_ports = []
    if port_path:
        port_path = Path(port_path)
        if port_path.exists():
            print(f"\n[2/4] 读取港口数据: {port_path.name}")
            if port_path.suffix.lower() in [".xlsx", ".xls"]:
                raw_ports = read_port_excel(port_path)
            elif port_path.suffix.lower() == ".pdf":
                raw_ports = read_port_pdf(port_path)
            print(f"  → 读取到 {len(raw_ports)} 条港口记录")
            
            # 从第一个港口提取默认登船口岸
            if raw_ports:
                first = raw_ports[0]
                default_port = first.get("_raw_port", "")
                default_joindate = first.get("_raw_arrival", "")
    
    # 3. 标准化
    print(f"\n[3/4] 标准化处理...")
    crew_normalized = normalize_crew(raw_crew, default_port, default_joindate)
    ports_normalized = normalize_ports(raw_ports) if raw_ports else []
    
    print(f"  → 船员清单: {len(crew_normalized)} 人")
    for c in crew_normalized:
        print(f"    [{c['序号']}] {c['姓名']} | {c['船员国籍 ']} | {c['船员职务']} | {c['证件类型']} | {c['登船口岸']}")
    print(f"  → 港口清单: {len(ports_normalized)} 条")
    for p in ports_normalized:
        print(f"    [{p['序号']}] {p['停靠港口']} | {p['国家/地区名称']} | {p['进港时间']} ~ {p['离港时间']}")
    
    # 4. 写入Excel
    print(f"\n[4/4] 生成输出文件...")
    wb = openpyxl.load_workbook(TEMPLATE_PATH)
    
    fill_crew_sheet(wb, crew_normalized)
    fill_goods_sheet(wb, crew_normalized)
    if ports_normalized:
        fill_port_sheet(wb, ports_normalized)
    
    # 保存
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    if not output_name:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_name = f"单证录入_{ts}"
    output_path = OUTPUT_DIR / f"{output_name}.xlsx"
    wb.save(output_path)
    print(f"\n✅ 完成！输出文件: {output_path}")
    return output_path


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("用法: python3 单证录入核心.py <crew文件> [port_of_call文件] [输出名]")
        print("示例: python3 单证录入核心.py crew.xlsx port.xlsx 2025航次报告")
        sys.exit(1)
    
    crew_path = sys.argv[1]
    port_path = sys.argv[2] if len(sys.argv) > 2 else None
    output_name = sys.argv[3] if len(sys.argv) > 3 else None
    
    result = process(crew_path, port_path, output_name)
    if not result:
        sys.exit(1)
